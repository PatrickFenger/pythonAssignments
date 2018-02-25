import webget
import webget
import csv
import openpyxl
import xlrd
import matplotlib.pyplot as plt
import re
#webget.download("https://ucr.fbi.gov/crime-in-the-u.s/2013/crime-in-the-u.s.-2013/tables/1tabledatadecoverviewpdf/table_1_crime_in_the_united_states_by_volume_and_rate_per_100000_inhabitants_1994-2013.xls/output.xls")
#url ="https://ucr.fbi.gov/crime-in-the-u.s/2013/crime-in-the-u.s.-2013/tables/1tabledatadecoverviewpdf/table_1_crime_in_the_united_states_by_volume_and_rate_per_100000_inhabitants_1994-2013.xls/output.xls"
#print(os.system("s"))

xl_workbook = xlrd.open_workbook("output.xls")

# List sheet names, and pull a sheet by name
#
sheet_names = xl_workbook.sheet_names()


xl_sheet = xl_workbook.sheet_by_name(sheet_names[0])
def change_in_crime():
    dict = {}
    for row_idx in range(4, 24):
        key = list(str(xl_sheet.cell(row_idx,0)).split(":"))[1]
        if "." in key:
            key = float(key)
        elif "\'" in key:
            temp_if = key.split("\'")
            print(temp_if)
            key = temp_if[1]
        sumd = 0.0
        for col_idx in range(2,20,2):
            temp = str(xl_sheet.cell(row_idx, col_idx)).split(":")
            sumd +=float(temp[1])
        print()
        dict[int(key)]=sumd

    listx = list(dict.keys())
    listt = list(dict.values())

    plt.bar(listx,listt,color='g')
    plt.xlabel('Year', fontsize=18)
    plt.ylabel('Amount of crime', fontsize=16)
    plt.show()

#print(int(float("1994.0")))

def most_common_type_of_crime():
    dict = {}
    for row_idx in range(4, 24):
        key = list(str(xl_sheet.cell(row_idx, 0)).split(":"))[1]
        if "." in key:
            key = float(key)
        elif "\'" in key:
            temp_if = key.split("\'")
            key = temp_if[1]
        sumd = 0.0
        most_common_crime = ""
        for col_idx in range(2, 20, 2):
            temp = float(list(str(xl_sheet.cell(row_idx, col_idx)).split(":"))[1])
            if sumd < temp:
                sumd = temp
                col_value = list(str(xl_sheet.cell(3, col_idx)).split(":"))
                most_common_crime = col_value[1].replace('\\n', ' ')
        dict[int(key)] = most_common_crime
    return dict

xl_workbook1 = xlrd.open_workbook("output1.xls")

# List sheet names, and pull a sheet by name
#
sheet_names1 = xl_workbook1.sheet_names()


xl_sheet1 = xl_workbook1.sheet_by_name(sheet_names1[0])

def crime_location():
     state=""
     dict={}
     dict_plot={}
     for row_idx in range(4, xl_sheet1.nrows):
        temp_state = list(str(xl_sheet1.cell(row_idx, 0)).split(":"))[1]
        city =  list(str(xl_sheet1.cell(row_idx, 1)).split(":"))[1]
        temp_list = []
        temp_dict = {}
        temp_dict[city]=[]
        temp_sum = 0.0
        if re.match("\'([A-Za-z]+)\'",temp_state):
            violent_crime_sum = 0.0
            murder_sum = 0.0
            rape1_sum = 0.0
            rap_legacy_sum = 0.0
            robber_sum = 0.0
            assault_sum=0.0
            property_sum = 0.0
            burglary_sum = 0.0
            larency_sum = 0.0
            m_vechiel_sum = 0.0
            arson_sum = 0.0

            state=temp_state
            dict[state]=[]
        for col in range(3,14):
            temp_value = list(str(xl_sheet1.cell(row_idx, col)).split(":"))[1]
            if col == 3:

                try:
                    state_sum += float(temp_value)
                    #temp_sum += float(temp_value)
                except ValueError:
                    print("not number")
        #dict_plot[state]=state_sum
        ##temp_dict[city].append(temp_sum)
        ##dict[state].append(temp_dict)
     print(dict_plot)

     return dict
crime_location()








